'''
------------------------------------------------------------------
This is a telegram bot use for store poshana tracker hight and weight into a excel file 
you just have to past your api token and its done.
------------------------------------------------------------------
'''
import telebot , os
import time
from telebot import apihelper
from openpyxl import load_workbook, Workbook
from colorama import Fore
import colorama
colorama.init(autoreset=True)

# Replace 'YOUR_API_TOKEN' with the actual API token provided by BotFather
API_TOKEN = 'YOUR_API_TOKEN'

bot = telebot.TeleBot(API_TOKEN)

def send_welcome(message):
    chat_id = message.chat.id
    welcome_message = "Welcome to the Copy Cat 🐈 Bot! 😊 Send me any text, and I'll copy it to your Pc clipboard.📋"
    bot.send_message(chat_id, welcome_message)

def save_to_excel(data):
    excel_file = 'Database\\poshana_data.xlsx'
    try:
        if os.path.exists(excel_file):
            workbook = load_workbook(excel_file)
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Child Name", "Mobile No", "Mother Name", "Height", "Weight", "Updating Date"])

        sheet = workbook.active
        sheet.append(data)
        workbook.save(excel_file)
        message = Fore.GREEN+" :Data saved to Excel.✅🐈‍⬛"
        child = f"{data[0]}"
        print(child+message)
    except Exception as e:
        print("Error saving data to Excel:", e)

def handle_data(message):
    chat_id = message.chat.id
    text = message.text
    lines = str(text).splitlines()
    num_lines = len(lines)
    first_name = lines[0]

    if num_lines == 6:
        data = [line.strip() for line in lines]
        save_to_excel(data)
        bot.send_message(chat_id, f"{first_name} Data saved to  Excel successfully✅🐈🎉")
    else:
       bot.send_message(chat_id, f"{first_name} Data Does not Matched with Format ❌")

@bot.message_handler(commands=['start'])
def handle_start(message):
    try:
        send_welcome(message)
    except Exception as e:
        print(f"An error occurred: {e}")

@bot.message_handler(func=lambda message: True)
def handle_text_message(message):
    try:
        handle_data(message)
    except Exception as e:
        print(f"An error occurred: {e}")

def banner():
    os.system("cls")
    print(Fore.CYAN+'''

╭━━━╮╱╱╱╱╱╭╮╱╱╱╱╱╱╱╱╭━━━━╮╱╱╱╱╱╱╭╮╱╱╱╱╱╱╱╭━━━╮╱╱╱╱╱╱╱╱╱╭╮
┃╭━╮┃╱╱╱╱╱┃┃╱╱╱╱╱╱╱╱┃╭╮╭╮┃╱╱╱╱╱╱┃┃╱╱╱╱╱╱╱┃╭━╮┃╱╱╱╱╱╱╱╱╭╯╰╮
┃╰━╯┣━━┳━━┫╰━┳━━┳━╮╱╰╯┃┃┣┻┳━━┳━━┫┃╭┳━━┳━╮┃┃╱┃┣━━┳━━┳┳━┻╮╭╋━━┳━╮╭━━┳━━╮
┃╭━━┫╭╮┃━━┫╭╮┃╭╮┃╭╮╮╱╱┃┃┃╭┫╭╮┃╭━┫╰╯┫┃━┫╭╯┃╰━╯┃━━┫━━╋┫━━┫┃┃╭╮┃╭╮┫╭━┫┃━┫
┃┃╱╱┃╰╯┣━━┃┃┃┃╭╮┃┃┃┃╱╱┃┃┃┃┃╭╮┃╰━┫╭╮┫┃━┫┃╱┃╭━╮┣━━┣━━┃┣━━┃╰┫╭╮┃┃┃┃╰━┫┃━┫
╰╯╱╱╰━━┻━━┻╯╰┻╯╰┻╯╰╯╱╱╰╯╰╯╰╯╰┻━━┻╯╰┻━━┻╯╱╰╯╱╰┻━━┻━━┻┻━━┻━┻╯╰┻╯╰┻━━┻━━╯  
                                
                        ʕ•́ᴥ•̀ʔっ
-------------------------------------------------------------
This Telegram bot is designed for Poshana Traker data entry.
Author - Rajkishor Patra
version - 1
-------------------------------------------------------------
          ''')

if __name__ == "__main__":
    while True:
        try:
            banner()
            bot.polling()
        except apihelper.ApiException as api_ex:
            if "Connection aborted" in str(api_ex):
                print("Connection error. Retrying in 10 seconds...")
                time.sleep(10)
            else:
                print(f"ApiException occurred: {api_ex}")
                break
        except Exception as e:
            print(f"An error occurred: {e}")
            break
